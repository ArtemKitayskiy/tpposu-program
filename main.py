import sys
import tkinter as tk
import datetime
import numpy as np
import openpyxl
import pandas as pd
from get_data import *
from tkinter import messagebox, ttk
from bdUtils import addDataToDB, signAndLog
from openpyxl.styles import Font


global user_login
user_login = ""
columns = tuple(conf.keys())

def out_table(window_name, data=None):

    # Создаем Treeview
    
    tree = ttk.Treeview(window_name, show="headings", columns=columns)
    for column in columns:
        tree.heading(f'{column}', text=column)
    # Устанавливаем заголовки
    # tree.heading('ID', text='ID')
    # tree.heading('Имя', text='Имя')
    # tree.heading('Возраст', text='Возраст')
    # tree.heading('Город', text='Город')
    if data:
        errs = data['errors']
        data = data['result']
    else:
        data = []

    for row in data:
        tree.insert('', tk.END, values=row)

    # Установим автоматическое масштабирование ширины колонок
    for column in tree['columns']:
        tree.column(column, width = 70,stretch=True)


    # Упаковываем Treeview
    tree.pack(expand=True, fill='both')


def on_key_press(event):
    if event.keysym == 'Return':
        button_login.invoke()

def center_window(window):
    window.update_idletasks()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    window_width = window.winfo_width()
    window_height = window.winfo_height()

    x_coordinate = (screen_width - window_width) // 2
    y_coordinate = (screen_height - window_height) // 2

    window.geometry(f"+{x_coordinate}+{y_coordinate}")

def registration_window():
    registration_window = tk.Toplevel(root)
    root.withdraw()
    registration_window.title("Окно регистрации")
    registration_window.geometry("400x200")
    center_window(registration_window)
    registration_window.resizable(False, False)

    label_username = tk.Label(registration_window, text="Логин:")
    label_username.grid(row=0, column=0, sticky=tk.E)
    label_username.place(x=100, y=50)

    reg_user = tk.Entry(registration_window)
    reg_user.grid(row=0, column=1)
    reg_user.place(x=150, y=50)

    label_password = tk.Label(registration_window, text="Пароль:")
    label_password.grid(row=1, column=0, sticky=tk.E)
    label_password.place(x=100, y=80)

    reg_password = tk.Entry(registration_window, show="*")
    reg_password.grid(row=1, column=1)
    reg_password.place(x=150, y=80)


    def add_user():

        username = reg_user.get()
        password = reg_password.get()
        if username =='' and password == '':
            messagebox.showerror("Ошибка регистарции", "Введите имя пользователя и пароль")
        elif username == '':
            messagebox.showerror("Ошибка регистарции", "Введите имя пользователя")
        elif password == '':
            messagebox.showerror("Ошибка регистарции", "Введите пароль")
        else:
            messagebox.showinfo('Информация о регистрации',signAndLog.add_user(username, password))
            registration_window.destroy()
            root.deiconify()
            


    button_login = tk.Button(registration_window, text="Регистрация", command=add_user, width=10, height=2)
    button_login.grid(row=2, columnspan=2)
    button_login.place(x=150, y=110)
    

def check_credentials():
    username = entry_username.get()
    password = entry_password.get()
    if username == '' and password == '':
        messagebox.showerror("Ошибка авторизации", "Введите имя пользователя и пароль")
    elif username == '':
        messagebox.showerror("Ошибка авторизации", "Введите имя пользователя")
    elif password == '':
        messagebox.showerror("Ошибка авторизации", "Введите пароль")
    else:
        result = signAndLog.check_user_password(username,password)

        if result[0] == True:
            messagebox.showinfo("Успешная авторизация", "Вы успешно авторизованы!")
            signAndLog.set_user_active_status(username)
            user_login = username
            root.withdraw()
            open_main_window()
        else:
            messagebox.showerror("Ошибка авторизации", 'Неверное имя пользователя или пароль')
    

def open_main_window():
    main_window = tk.Toplevel(root)
    main_window.title("Главное окно")
    main_window.geometry("400x200")
    main_window.resizable(False, False)
    center_window(main_window)

    # Создаем словарь, где ключами являются названия кнопок, а значениями - функции для открытия соответствующих окон
    buttons_actions = {
        "Регистрация данных": open_registration_window,
        "Проведение HTP": open_htp_window,
        "Управление данными": open_data_management_window,
        "Информация о ПО": open_software_info_window,
        "Выход": sys.exit,
    }

    # Создаем кнопки на основе элементов словаря
    for text, action in buttons_actions.items():
        button = tk.Button(main_window, text=text, command=action)
        button.pack(pady=5)


def open_registration_window():

    new_registration_window = tk.Toplevel(root)
    new_registration_window.title("Окно регистрации данных")
    new_registration_window.geometry("1280x720")
    center_window(new_registration_window)
    data = []
    columns = tuple(conf.keys())
    tree = ttk.Treeview(new_registration_window, show="headings", columns=columns, height=20)
    def out_data():
        nonlocal comm_label
        nonlocal tree
        nonlocal columns
        nonlocal data
        data = 0
        data = []
        count = int(frames_entry.get())

	    # Очистка всех элементов в Treeview
        for item in tree.get_children():
            tree.delete(item)

        for i in range(0,count):
            data.append([i+1]+get_data()['result'])
            
            errors = get_data()['errors']
            if len(errors )!=0:
                messagebox.showerror("Ошибка", errors)
          
        for column in columns:
        	tree.heading(f'{column}', text=column)

        for row in data:
        	tree.insert('', tk.END, values=row)

    	# Установим автоматическое масштабирование ширины колонок
        for column in tree['columns']:
        	tree.column(column, width = 70,stretch=True)

    	# Упаковываем Treeview
        # tree.pack(expand=True, fill='both')
        tree.grid(row=15, column=0, columnspan=10, padx=10, pady=10)
        scrollbar = ttk.Scrollbar(new_registration_window, orient="vertical", command=tree.yview)
        scrollbar.grid(row=15, column=10, sticky="ns")
        tree.configure(yscrollcommand=scrollbar.set)

    def safe_data():
        nonlocal data
        if data !=[]:
            user_login = signAndLog.check_signed_users()
            user_id = signAndLog.get_userid_by_login(user_login)
            addDataToDB.add_data_in_table(data, user_id, datetime.datetime.now(), comm_entry.get())
        else:
            messagebox.showerror("Ошибка", "Сначала необходимо провести опрос")
        

    # Создание и размещение виджетов
    frames_label = tk.Label(new_registration_window, text="Число кадров:", width=15)
    frames_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")

    frames_entry = tk.Entry(new_registration_window, width=15)
    frames_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    start_button = tk.Button(new_registration_window, text="Начать опрос", command=out_data, width = 15)
    start_button.grid(row=0, column=2, padx=5, pady=5, sticky="w")

    comm_label = tk.Label(new_registration_window, text="Комментарий:", width=15)
    comm_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")

    comm_entry = tk.Entry(new_registration_window, width=15)
    comm_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    save_button = tk.Button(new_registration_window, text="Сохранить", command=safe_data, width = 15)
    save_button.grid(row=1, column=2, padx=5, pady=5, sticky="w")




    


def open_htp_window():
    new_htp_window = tk.Toplevel(root)
    new_htp_window.title("Окно проведения HTP")
    new_htp_window.geometry("1280x720")
    center_window(new_htp_window)
    tree_of_htp = ttk.Treeview(new_htp_window, columns=columns, show="headings", height=20)

    # Устанавливаем заголовки для колонок
    for col in columns:
        tree_of_htp.heading(col, text=col)
    # Установим автоматическое масштабирование ширины колонок
    for column in tree_of_htp['columns']:
        tree_of_htp.column(column, width = 80, stretch=True)
    # Заполните таблицу данными, если это необходимо
    # Пример добавления данных
    # Разместите таблицу внизу окна
    tree_of_htp.grid(row=0, column=0, columnspan=10, padx=10, pady=10, sticky='w')

    scrollbar = ttk.Scrollbar(new_htp_window, orient="vertical", command=tree_of_htp.yview)
    scrollbar.grid(row=0, column=10, sticky="ns")
    tree_of_htp.configure(yscrollcommand=scrollbar.set)

    info_text_htp = tk.Text(new_htp_window, wrap="word", width=40, height=13)
    
    info_text_htp.grid(row=18, column=0, padx=10, pady=10, rowspan=3, sticky='w')

    def get_excel(file_name):
        nonlocal info_text_htp
        try:
            # Load the xlsx file
            excel_data = pd.read_excel(f'{file_name}.xlsx')
            # Read the values of the file in the dataframe
            data = pd.DataFrame(excel_data)
            info = "\n".join(list(data['Отчет о регистрации данных с объекта управления'])[1:4])
            data = data.iloc[6:].values
            info_text_htp.delete('1.0', tk.END)
            for item in tree_of_htp.get_children():
                tree_of_htp.delete(item)
            for row in data:
                row = list(row)
                tree_of_htp.insert('', tk.END, values=row)

            info_text_htp.insert(tk.END, info)
            
                    
        except:
            messagebox.showerror("Ошибка имени файла", "Такого файла не существует")

    file_label_name = tk.Label(new_htp_window, text="Введите имя файла:")
    file_label_name.grid(row=19, column=8, padx=10, pady=5)

    entry_name = tk.Entry(new_htp_window)
    entry_name.grid(row=19, column=9, padx=10, pady=5)

    button_load_xls = tk.Button(new_htp_window, text="Загрузить из excel", width=20, height=3, command = lambda :get_excel(entry_name.get()))
    button_load_xls.grid(row=20, column=9, padx=10, pady=5, rowspan=2)


def open_data_management_window():
    selected_date = None
    data = []
    column_means = []
    user_id = 0
    new_data_management_window = tk.Toplevel(root)
    new_data_management_window.title("Окно управления данными")
    new_data_management_window.geometry("1100x850")
    center_window(new_data_management_window)

    columns = tuple(conf.keys())
    tree = ttk.Treeview(new_data_management_window, columns=columns, show="headings", height=20)

    # Устанавливаем заголовки для колонок
    for col in columns:
        tree.heading(col, text=col)
    # Установим автоматическое масштабирование ширины колонок
    for column in tree['columns']:
        tree.column(column, width = 80, stretch=True)
    # Заполните таблицу данными, если это необходимо
    # Пример добавления данных:
    # tree.insert("", "end", values=("1", "John", "30", "New York"))

    # Разместите таблицу внизу окна
    tree.grid(row=2, column=0, columnspan=10, padx=10, pady=10, sticky='w')

    # Добавьте прокрутку для таблицы при необходимости
    scrollbar = ttk.Scrollbar(new_data_management_window, orient="vertical", command=tree.yview)
    scrollbar.grid(row=2, column=9, sticky="ns")
    tree.configure(yscrollcommand=scrollbar.set)

    tree_of_mean = ttk.Treeview(new_data_management_window, columns=columns[:-2], show="headings", height=1)

    # Устанавливаем заголовки для колонок
    for col in columns[:-2]:
        tree_of_mean.heading(col, text=col)
    # Установим автоматическое масштабирование ширины колонок
    for column in tree_of_mean['columns']:
        tree_of_mean.column(column, width = 80, stretch=True)
    # Заполните таблицу данными, если это необходимо
    # Пример добавления данных
    # Разместите таблицу внизу окна
    tree_of_mean.grid(row=9, column=0, columnspan=10, padx=10, pady=10, sticky='w')
    info_text = tk.Text(new_data_management_window, wrap="word", width=40, height=13)
    info_text.grid(row=10, column=0, padx=10, pady=10, rowspan=3, sticky='w')
        


    def combobox_changed(event):

        def date_changed(event):
            nonlocal user_id
            nonlocal selected_date
            nonlocal data
            nonlocal column_means
            # Очистка всех элементов в Treeview
            for item in tree.get_children():
                tree.delete(item)
            info_text.delete('1.0', tk.END)

            selected_date = dropdown_3.get()
            data = addDataToDB.get_exp_by_date(selected_date)

            for row in data:
                tree.insert('', tk.END, values=row)
            # Преобразование данных в массив NumPy
            data_array = np.array(data)[:, :-2].astype(float)

            # Расчет среднего по каждому столбцу
            column_means = np.mean(data_array, axis=0)
            column_means = np.round(column_means,2)
            means = {
                "Средние значения": 0,
                "Cреднее для Канала 1": 0,
                "Cреднее для Канала 2": 0,
                "Cреднее для Канала 3": 0,
                "Cреднее для Канала 4": 0,
                "Cреднее для Канала 6 сред.": 0,
                "Cреднее для Канала 6 дисп.": 0,
                "Cреднее для Канала 14": 0,
                "Cреднее для Канала 44": 0,
                "Cреднее для Канала 54": 0,
                "Cреднее для Канала 65": 0,
            }
            means  = dict(zip(means, column_means))
            means["Средние значения"] = "Средние значения"
            s = 'Комментарий:'+addDataToDB.get_exp_comment(selected_date)+'\n'
            for i,v in means.items():
                if i!="Средние значения":
                    s += f'{i}:{v}\n'
            # Добавляем текст в виджет Text
            info_text.insert(tk.END, s)

            tree_of_mean.insert('', tk.END, values=tuple(means.values()))



        selected_value = dropdown.get()

        user_id = signAndLog.get_userid_by_login(selected_value)

        dates = addDataToDB.get_exp_data_by_id(user_id)
        dates = [i[0] for i in dates]
        dropdown_frame_3 = tk.Frame(new_data_management_window)
        dropdown_frame_3.grid(row=1, column=0, padx=10, pady=10)

        label_3 = tk.Label(dropdown_frame_3, text="Дата")
        label_3.pack(side='top', anchor='w')

        dropdown_values_3 = ["Элемент X", "Элемент Y", "Элемент Z"]
        dropdown_3 = ttk.Combobox(dropdown_frame_3, values=dates, state="readonly")
        dropdown_3.pack(side='left')
        dropdown_3.bind("<<ComboboxSelected>>", date_changed)

    dropdown_frame = tk.Frame(new_data_management_window)
    dropdown_frame.grid(row=0, column=0, padx=10, pady=10)

    label = tk.Label(dropdown_frame, text="Ответственный")
    label.pack(side='top', anchor='w')

    dropdown_values = signAndLog.get_users_logins()
    dropdown = ttk.Combobox(dropdown_frame, values=dropdown_values, state='readonly')
    dropdown.pack(side='left')
    dropdown.bind("<<ComboboxSelected>>", combobox_changed)

    dropdown_frame_2 = tk.Frame(new_data_management_window)
    dropdown_frame_2.grid(row=0, column=1, padx=10, pady=10)

    label_2 = tk.Label(dropdown_frame_2, text="Столбец для сортировки")
    label_2.pack(side='top', anchor='w')

    dropdown_values_2 = ["Элемент A", "Элемент B", "Элемент C", "Элемент D"]
    dropdown_2 = ttk.Combobox(dropdown_frame_2, values=columns, state="readonly")
    dropdown_2.set(columns[0])
    dropdown_2.pack(side='left')


    dropdown_frame_4 = tk.Frame(new_data_management_window)
    dropdown_frame_4.grid(row=1, column=1, padx=10, pady=10)

    label_4 = tk.Label(dropdown_frame_4, text="Порядок сортировки")
    label_4.pack(side='top', anchor='w')

    dropdown_values_4 = ["По возрастанию", "По убыванию"]
    dropdown_4 = ttk.Combobox(dropdown_frame_4, values=dropdown_values_4, state="readonly")
    dropdown_4.set("По возрастанию")
    dropdown_4.pack(side='left')

    dropdown_frame_5 = tk.Frame(new_data_management_window)
    dropdown_frame_5.grid(row=0, column=2, padx=10, pady=10)

    label_5 = tk.Label(dropdown_frame_5, text="Канал")
    label_5.pack(side='top', anchor='w')

    dropdown_values_5 = ["Элемент Q", "Элемент R", "Элемент S"]
    dropdown_5 = ttk.Combobox(dropdown_frame_5, values=columns, state="readonly")
    dropdown_5.set(columns[0])
    dropdown_5.pack(side='left')

    min_entry_frame = tk.Frame(new_data_management_window)
    min_entry_frame.grid(row=1, column=2, padx=10, pady=10)

    label_6 = tk.Label(min_entry_frame, text="Минимальное")
    label_6.pack(side='top', anchor='w')

    entry_1 = tk.Entry(min_entry_frame)
    entry_1.pack(side='top')

    man_entry_frame = tk.Frame(new_data_management_window)
    man_entry_frame.grid(row=1, column=3, padx=10, pady=10)

    label_7 = tk.Label(man_entry_frame, text="Максимальное")
    label_7.pack(side='top', anchor='w')

    entry_2 = tk.Entry(man_entry_frame)
    entry_2.pack(side='top')

    def get_filtered_data(col, sort_type, min, max):
        nonlocal data
        for item in tree.get_children():
            tree.delete(item)
        sort_type = 'DESC' if sort_type == 'По убыванию' else ''

        if min!='' and max!='':
            filter_condition = f'AND {conf[col]} BETWEEN {min} AND {max}'
        elif min!='':
            filter_condition = f'AND {conf[col]} >= {min}'
        elif max!='':
            filter_condition = f'AND {conf[col]} <= {max}'
        else:
            filter_condition = ''
        
        data = addDataToDB.get_sorted_data(selected_date, conf[col], sort_type, filter_condition)
        print(data)
        if data != []:
        	data_array = np.array(data)[:, :-2].astype(float)

        	# Расчет среднего по каждому столбцу
        	column_means = np.mean(data_array, axis=0)
        	column_means = np.round(column_means,2)
        	means = {
            	"Cреднее для Канала 1": 0,
            	"Cреднее для Канала 2": 0,
            	"Cреднее для Канала 3": 0,
            	"Cреднее для Канала 4": 0,
            	"Cреднее для Канала 6 сред.": 0,
            	"Cреднее для Канала 6 дисп.": 0,
            	"Cреднее для Канала 14": 0,
            	"Cреднее для Канала 44": 0,
            	"Cреднее для Канала 54": 0,
            	"Cреднее для Канала 65": 0,
        	}
        	means  = dict(zip(means, column_means[1:]))
        	s = 'Комментарий:'+addDataToDB.get_exp_comment(selected_date)+'\n'
        	for i,v in means.items():
            		s += f'{i}:{v}\n'
        	# Добавляем текст в виджет Text
        	info_text.insert(tk.END, s)
        else:
        	s = 'Комментарий:'+addDataToDB.get_exp_comment(selected_date)+'\n'
        	info_text.insert(tk.END, s)

        for row in data:
                tree.insert('', tk.END, values=row)

    button = tk.Button(new_data_management_window, text="Установить фильтр", width=20, height=3, command = lambda: get_filtered_data(dropdown_5.get(), dropdown_4.get(), entry_1.get(), entry_2.get()))
    button.grid(row=0, column=4, padx=10, pady=10, rowspan=2)

    def data_to_excel(name):
        if name == "":
                messagebox.showerror("Ошибка имени файла", "Ошибка имени файла")
        df = pd.DataFrame(data, columns=columns)
        # Запись DataFrame в файл Excel
        df.to_excel(f'{name}.xlsx', index=False)
        # обработка excel файла
        wb_obj = openpyxl.load_workbook(f'{name}.xlsx')
        sheet_obj = wb_obj.active
        sheet_obj.insert_rows(1, 6)
        
        sheet_obj.cell(row = 1, column = 1).value = "Отчет о регистрации данных с объекта управления"
        sheet_obj.cell(row = 1, column = 1).font = Font(bold=True)

        sheet_obj.cell(row = 3, column = 1).value = f"Пользователь: {signAndLog.check_signed_users()}"

        sheet_obj.cell(row = 4, column = 1).value = f"Дата регистрации: {selected_date}"
        
        comment = addDataToDB.get_exp_comment(selected_date)
        if len(comment)==0:
            comment = "Нет комментария"
        sheet_obj.cell(row = 5, column = 1).value = f"Комментарий: {comment}"
        wb_obj.save(f'{name}.xlsx')

        messagebox.showinfo('Запись успешна','Запись успешна')

    def excel_window():
        nonlocal data
        if data == []:
            messagebox.showerror("Ошибка", "Сначала выберите эксперимент")
        else:
            excel_window = tk.Toplevel(root)
            excel_window.title("Окно записи")
            excel_window.geometry("300x100")
            center_window(excel_window)
            excel_window.resizable(False, False)

            label_name = tk.Label(excel_window, text="Введите имя файла:")
            label_name.grid(row=0, column=0, padx=10, pady=10)

            entry_name = tk.Entry(excel_window)
            entry_name.grid(row=1, column=0, padx=10, pady=10)
            button_3 = tk.Button(excel_window, text="Записать в excel", width=20, height=3, command = lambda: data_to_excel(entry_name.get()))
            button_3.grid(row=0, column=1, padx=10, pady=10, rowspan=2)



    button_2 = tk.Button(new_data_management_window, text="Записать в excel", width=20, height=3, command = excel_window)
    button_2.grid(row=10, column=4, padx=10, pady=10, rowspan=2)

    



def open_software_info_window():
    new_software_info_window = tk.Toplevel(root)
    new_software_info_window.title("Окно информации о ПО")
    new_software_info_window.geometry("400x200")
    center_window(new_software_info_window)
    label = tk.Label(new_software_info_window, text="Текст")
    label.pack()


root = tk.Tk()
root.title("Окно авторизации")
root.geometry("400x200")
root.resizable(False, False)
center_window(root)

# Создание и размещение виджетов
label_username = tk.Label(root, text="Логин:")
label_username.grid(row=0, column=0, sticky=tk.E)
label_username.place(x=100, y=50)

entry_username = tk.Entry(root)
entry_username.grid(row=0, column=1)
entry_username.place(x=150, y=50)

label_password = tk.Label(root, text="Пароль:")
label_password.grid(row=1, column=0, sticky=tk.E)
label_password.place(x=100, y=80)

entry_password = tk.Entry(root, show="*")
entry_password.grid(row=1, column=1)
entry_password.place(x=150, y=80)

button_login = tk.Button(root, text="Ввод", command=check_credentials, width=10, height=2)
button_login.grid(row=2, columnspan=2)
button_login.place(x=150, y=110)

button_registraion = tk.Button(root, text="Регистрация", command=registration_window, width=10, height=2)
button_registraion.grid(row=2, columnspan=2)
button_registraion.place(x=150, y=150)


root.bind("<Return>", on_key_press)
# Запуск главного цикла обработки событий
root.mainloop()
